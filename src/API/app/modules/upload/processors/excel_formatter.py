"""Excel formatter for environmental analysis output with conditional formatting."""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Optional

from app.core.logger import get_logger

logger = get_logger(__name__)


class ExcelFormatter:
    """Service for formatting CSV output to Excel with conditional formatting."""
    
    def __init__(self):
        # Define colors for conditional formatting
        self.red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
        self.green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        
        # Define fonts
        self.header_font = Font(color="FFFFFF", bold=True, size=11)  # White, bold
        self.data_font = Font(size=10)
        
        # Define alignment
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Define borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define significance columns to format
        self.significance_columns = [
            "Vegetation (NDVI)-Significance",
            "Built-up Area (NDBI)-Significance", 
            "Water/Moisture (NDWI)-Significance"
        ]
    
    def convert_csv_to_formatted_xlsx(self, csv_path: Path, output_path: Optional[Path] = None) -> Path:
        """
        Convert CSV to formatted XLSX with conditional formatting.
        
        Args:
            csv_path: Path to input CSV file
            output_path: Path for output XLSX file (optional)
            
        Returns:
            Path to the created XLSX file
        """
        try:
            logger.info(f"Converting CSV to formatted XLSX: {csv_path}")
            
            # Read CSV file
            df = pd.read_csv(csv_path)
            
            # Generate output path if not provided
            if output_path is None:
                output_path = csv_path.with_suffix('.xlsx')
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Environmental Analysis"
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format the worksheet
            self._format_worksheet(ws, df)
            
            # Apply conditional formatting to significance columns
            self._apply_conditional_formatting(ws, df)
            
            # Auto-adjust column widths
            self._adjust_column_widths(ws)
            
            # Save workbook
            wb.save(output_path)
            
            logger.info(f"Formatted XLSX created: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to convert CSV to formatted XLSX: {str(e)}")
            raise
    
    def _format_worksheet(self, ws, df):
        """Apply general formatting to the worksheet."""
        try:
            # Format header row
            for col_num, column in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
            
            # Format data rows
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = self.data_font
                    cell.border = self.thin_border
                    
                    # Center align numeric columns, left align text columns
                    column_name = df.columns[col_num - 1]
                    if df[column_name].dtype in ['int64', 'float64']:
                        cell.alignment = self.center_alignment
                    else:
                        cell.alignment = self.left_alignment
            
            logger.debug("General worksheet formatting applied")
            
        except Exception as e:
            logger.error(f"Failed to format worksheet: {str(e)}")
            raise
    
    def _apply_conditional_formatting(self, ws, df):
        """Apply conditional formatting to significance columns."""
        try:
            # Find significance column indices
            significance_col_indices = {}
            for col_name in self.significance_columns:
                if col_name in df.columns:
                    col_index = df.columns.get_loc(col_name) + 1  # Excel is 1-indexed
                    significance_col_indices[col_name] = col_index
            
            logger.info(f"Found significance columns: {list(significance_col_indices.keys())}")
            
            # Apply conditional formatting to each significance column
            for col_name, col_index in significance_col_indices.items():
                self._format_significance_column(ws, df, col_name, col_index)
            
            logger.debug("Conditional formatting applied to significance columns")
            
        except Exception as e:
            logger.error(f"Failed to apply conditional formatting: {str(e)}")
            raise
    
    def _format_significance_column(self, ws, df, col_name, col_index):
        """Format a specific significance column with conditional colors."""
        try:
            # Get the column data
            col_data = df[col_name]
            
            # Apply formatting to each cell in the column (skip header)
            for row_num in range(2, len(col_data) + 2):
                cell = ws.cell(row=row_num, column=col_index)
                cell_value = str(cell.value).strip().lower() if cell.value else ""
                
                # Apply conditional formatting based on value
                if cell_value == "yes":
                    cell.fill = self.red_fill
                    logger.debug(f"Applied red formatting to {col_name} row {row_num}: {cell_value}")
                elif cell_value == "no":
                    cell.fill = self.green_fill
                    logger.debug(f"Applied green formatting to {col_name} row {row_num}: {cell_value}")
                
                # Ensure text is centered
                cell.alignment = self.center_alignment
            
            logger.info(f"Formatted significance column: {col_name}")
            
        except Exception as e:
            logger.error(f"Failed to format significance column {col_name}: {str(e)}")
            raise
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths for better readability."""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with some padding
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            logger.debug("Column widths adjusted")
            
        except Exception as e:
            logger.error(f"Failed to adjust column widths: {str(e)}")
            # Don't raise exception for column width adjustment failures
    
    def create_summary_sheet(self, wb, df):
        """Create a summary sheet with statistics."""
        try:
            # Create summary worksheet
            summary_ws = wb.create_sheet("Summary")
            
            # Add title
            summary_ws['A1'] = "Environmental Analysis Summary"
            summary_ws['A1'].font = Font(size=16, bold=True)
            summary_ws['A1'].alignment = self.center_alignment
            
            # Merge cells for title
            summary_ws.merge_cells('A1:D1')
            
            # Add statistics
            row = 3
            
            # Total records
            summary_ws[f'A{row}'] = "Total Records:"
            summary_ws[f'B{row}'] = len(df)
            row += 1
            
            # Significance statistics for each environmental index
            for col_name in self.significance_columns:
                if col_name in df.columns:
                    col_data = df[col_name]
                    yes_count = (col_data.str.lower() == 'yes').sum()
                    no_count = (col_data.str.lower() == 'no').sum()
                    
                    # Add section header
                    summary_ws[f'A{row}'] = col_name.replace('-Significance', '') + ":"
                    summary_ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Add counts
                    summary_ws[f'B{row}'] = "Significant (Yes):"
                    summary_ws[f'C{row}'] = yes_count
                    summary_ws[f'C{row}'].fill = self.red_fill
                    row += 1
                    
                    summary_ws[f'B{row}'] = "Not Significant (No):"
                    summary_ws[f'C{row}'] = no_count
                    summary_ws[f'C{row}'].fill = self.green_fill
                    row += 2
            
            # Format summary sheet
            for row_cells in summary_ws.iter_rows():
                for cell in row_cells:
                    cell.border = self.thin_border
                    cell.alignment = self.left_alignment
            
            # Adjust column widths
            for column in summary_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 30)
                summary_ws.column_dimensions[column_letter].width = adjusted_width
            
            logger.info("Summary sheet created")
            
        except Exception as e:
            logger.error(f"Failed to create summary sheet: {str(e)}")
            # Don't raise exception for summary sheet creation failures


def format_environmental_analysis_excel(csv_path: Path, output_path: Optional[Path] = None) -> Path:
    """
    Convenience function to format environmental analysis CSV to Excel.
    
    Args:
        csv_path: Path to input CSV file
        output_path: Path for output XLSX file (optional)
        
    Returns:
        Path to the created XLSX file
    """
    formatter = ExcelFormatter()
    return formatter.convert_csv_to_formatted_xlsx(csv_path, output_path)


# Example usage
if __name__ == "__main__":
    # Test the formatter
    import sys
    
    if len(sys.argv) > 1:
        csv_file = Path(sys.argv[1])
        if csv_file.exists():
            output_file = format_environmental_analysis_excel(csv_file)
            print(f"Formatted Excel file created: {output_file}")
        else:
            print(f"CSV file not found: {csv_file}")
    else:
        print("Usage: python excel_formatter.py <csv_file_path>")