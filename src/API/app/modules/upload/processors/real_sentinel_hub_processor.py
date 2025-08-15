"""
Real Sentinel Hub API Integration for Geospatial Analysis
Replaces synthetic data generation with actual satellite imagery analysis
"""

import pandas as pd
import numpy as np
import yaml
import json
import time
from pathlib import Path
from typing import List, Dict, Any, Tuple
from datetime import datetime, timedelta
import logging

from sentinelhub import (
    SHConfig, 
    BBox, 
    CRS, 
    SentinelHubRequest, 
    DataCollection, 
    MimeType,
    bbox_to_dimensions
)

from app.core.exceptions import FileProcessingException
from app.core.logger import get_logger
from app.services.api_usage_service import APIUsageService

logger = get_logger(__name__)


class RealSentinelHubProcessor:
    """Real Sentinel Hub API processor for satellite imagery analysis."""
    
    def __init__(self, config_path: str = "sentinel_hub_config.yml", 
                 user_config_path: str = "sentinel_hub_user_config.yaml"):
        """Initialize the real Sentinel Hub processor."""
        
        # Load configurations
        self.config_data = self._load_config(config_path)
        self.user_config = self._load_user_config(user_config_path)
        
        # Setup Sentinel Hub configuration
        self.config = SHConfig()
        self.config.sh_client_id = self.config_data['sentinel_hub']['client_id']
        self.config.sh_client_secret = self.config_data['sentinel_hub']['client_secret']
        self.config.use_oauth = True
        
        # Processing parameters
        self.resolution = self.config_data['image_processing']['resolution']
        self.max_cloud_coverage = self.config_data['image_processing']['max_cloud_coverage']
        
        # Change detection thresholds
        self.thresholds = self.config_data['change_detection']
        
        logger.info("Real Sentinel Hub processor initialized")
        logger.info(f"Resolution: {self.resolution}m, Max cloud coverage: {self.max_cloud_coverage}%")
        logger.info(f"Sentinel Hub Client ID: {self.config.sh_client_id[:8]}...")
        
        # Initialize API call tracking
        self.api_call_count = 0
        self.total_api_time = 0.0
    
    def _load_config(self, config_path: str) -> Dict:
        """Load main configuration from YAML file."""
        try:
            with open(config_path, 'r') as file:
                return yaml.safe_load(file)
        except Exception as e:
            logger.error(f"Failed to load config from {config_path}: {str(e)}")
            raise FileProcessingException(f"Configuration loading failed: {str(e)}")
    
    def _load_user_config(self, user_config_path: str) -> Dict:
        """Load user configuration from YAML file."""
        try:
            with open(user_config_path, 'r') as file:
                return yaml.safe_load(file)
        except Exception as e:
            logger.warning(f"Failed to load user config from {user_config_path}: {str(e)}")
            return {}  # Return empty dict if user config is not available
    
    async def process_file(
        self,
        input_path: Path,
        output_dir: Path,
        dates: List[str],
        engagement_name: str,
        user_id: int,
        db_session = None
    ) -> Path:
        """Process geospatial data with real Sentinel Hub API calls with API limit checking."""
        
        logger.info(f"Starting real Sentinel Hub processing: {input_path}")
        
        try:
            # Step 1: Load the input file
            df = await self._load_file(input_path)
            
            # Step 2: Check API limits before processing
            if db_session:
                api_service = APIUsageService(db_session)
                required_calls = len(df) * 2  # 2 API calls per property (before + after periods)
                
                can_make_calls, error_message, usage_info = await api_service.check_api_limit(user_id, required_calls)
                
                if not can_make_calls:
                    logger.error(f"API limit exceeded for user {user_id}: {error_message}")
                    raise FileProcessingException(f"API Limit Exceeded: {error_message}")
                
                logger.info(f"API limit check passed for user {user_id}: {required_calls} calls needed, {usage_info['remaining_calls']} remaining")
            
            # Step 3: Add temporal period columns
            processed_df = await self._add_temporal_periods(df, dates)
            
            # Step 4: Get real satellite imagery indices using Sentinel Hub API
            processed_df, successful_calls = await self._get_real_satellite_indices(processed_df, dates)
            
            # Step 5: Update API usage counter for successful calls
            if db_session and successful_calls > 0:
                await api_service.increment_api_usage(user_id, successful_calls)
                logger.info(f"Updated API usage for user {user_id}: +{successful_calls} successful calls")
            
            # Step 6: Calculate differences and interpretations
            processed_df = await self._calculate_differences_and_interpretations(processed_df)
            
            # Step 7: Generate output file
            output_path = await self._generate_output_file(
                processed_df, 
                output_dir, 
                input_path.stem,
                engagement_name
            )
            
            logger.info(f"Real Sentinel Hub processing completed: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Real Sentinel Hub processing failed for {input_path}: {str(e)}")
            raise FileProcessingException(f"Sentinel Hub processing failed: {str(e)}")
    
    async def _load_file(self, file_path: Path) -> pd.DataFrame:
        """Load file into pandas DataFrame."""
        
        try:
            file_extension = file_path.suffix.lower()
            
            if file_extension == '.csv':
                df = pd.read_csv(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path)
            else:
                raise FileProcessingException(f"Unsupported file format: {file_extension}")
            
            if df.empty:
                raise FileProcessingException("File is empty or contains no data")
            
            logger.info(f"File loaded successfully: {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except Exception as e:
            raise FileProcessingException(f"Failed to load file: {str(e)}")
    
    async def _add_temporal_periods(self, df: pd.DataFrame, dates: List[str]) -> pd.DataFrame:
        """Add temporal period columns based on provided dates."""
        
        try:
            processed_df = df.copy()
            
            # Add temporal period columns
            processed_df['Before Period Start'] = dates[0] if len(dates) > 0 else '2022-11-01'
            processed_df['Before Period End'] = dates[1] if len(dates) > 1 else '2023-01-31'
            processed_df['After Period Start'] = dates[2] if len(dates) > 2 else '2025-01-01'
            processed_df['After Period End'] = dates[3] if len(dates) > 3 else '2025-03-31'
            
            return processed_df
            
        except Exception as e:
            raise FileProcessingException(f"Failed to add temporal periods: {str(e)}")
    
    async def _get_real_satellite_indices(self, df: pd.DataFrame, dates: List[str]) -> Tuple[pd.DataFrame, int]:
        """Get real satellite imagery indices using Sentinel Hub API.
        
        Returns:
            Tuple of (processed_dataframe, successful_api_calls_count)
        """
        
        try:
            processed_df = df.copy()
            
            # Define time periods
            before_period = (dates[0], dates[1]) if len(dates) >= 2 else ('2022-11-01', '2023-01-31')
            after_period = (dates[2], dates[3]) if len(dates) >= 4 else ('2025-01-01', '2025-03-31')
            
            logger.info(f"Analyzing {len(df)} properties with Sentinel Hub API")
            logger.info(f"Before period: {before_period}")
            logger.info(f"After period: {after_period}")
            
            # Initialize counters for batch summary
            successful_calls = 0
            failed_calls = 0
            successful_properties = []
            failed_properties = []
            
            # Process each property
            for index, row in processed_df.iterrows():
                lat = row.get('LATITUDE', 0)
                lon = row.get('LONGITUDE', 0)
                extent_ac = row.get('extent_ac', 0)
                
                logger.info(f"🏠 Processing property {index + 1}/{len(df)}: {lat:.6f}, {lon:.6f}")
                if extent_ac > 0:
                    logger.info(f"   📏 Land Area: {extent_ac:.2f} acres ({extent_ac * 4046.86:.0f} sq meters)")
                
                try:
                    # Get satellite data for before period
                    logger.info(f"   📅 Analyzing BEFORE period: {before_period}")
                    before_indices, before_error = await self._analyze_single_property(lat, lon, extent_ac, before_period)
                    
                    # Get satellite data for after period  
                    logger.info(f"   📅 Analyzing AFTER period: {after_period}")
                    after_indices, after_error = await self._analyze_single_property(lat, lon, extent_ac, after_period)
                    
                    # Check if both calls were successful (no error messages and not all zeros)
                    before_success = not before_error and any(v != 0.0 for v in before_indices.values())
                    after_success = not after_error and any(v != 0.0 for v in after_indices.values())
                    
                    if before_success and after_success:
                        successful_calls += 2
                        successful_properties.append(index)
                        logger.info(f"   ✅ Property {index + 1} analysis completed successfully")
                        
                        # Store results in DataFrame for successful properties
                        processed_df.at[index, 'Vegetation (NDVI)-Before Value'] = before_indices.get('ndvi', 0.0)
                        processed_df.at[index, 'Vegetation (NDVI)-After Value'] = after_indices.get('ndvi', 0.0)
                        
                        processed_df.at[index, 'Built-up Area (NDBI)-Before Value'] = before_indices.get('ndbi', 0.0)
                        processed_df.at[index, 'Built-up Area (NDBI)-After Value'] = after_indices.get('ndbi', 0.0)
                        
                        processed_df.at[index, 'Water/Moisture (NDWI)-Before Value'] = before_indices.get('ndwi', 0.0)
                        processed_df.at[index, 'Water/Moisture (NDWI)-After Value'] = after_indices.get('ndwi', 0.0)
                        
                        # Store success status for later (will be moved to end)
                        processed_df.at[index, '_temp_conversion_status'] = 'Successful'
                        
                    else:
                        failed_calls += (0 if before_success else 1) + (0 if after_success else 1)
                        successful_calls += (1 if before_success else 0) + (1 if after_success else 0)
                        failed_properties.append(index)
                        logger.warning(f"   ⚠️ Property {index + 1} analysis failed - INCLUDING in output with error message")
                        logger.warning(f"      Before success: {before_success}, After success: {after_success}")
                        
                        # Store zeros for failed properties (keep original input data intact)
                        processed_df.at[index, 'Vegetation (NDVI)-Before Value'] = 0.0
                        processed_df.at[index, 'Vegetation (NDVI)-After Value'] = 0.0
                        processed_df.at[index, 'Built-up Area (NDBI)-Before Value'] = 0.0
                        processed_df.at[index, 'Built-up Area (NDBI)-After Value'] = 0.0
                        processed_df.at[index, 'Water/Moisture (NDWI)-Before Value'] = 0.0
                        processed_df.at[index, 'Water/Moisture (NDWI)-After Value'] = 0.0
                        
                        # Store error message for later (will be moved to end)
                        error_message = before_error or after_error or "API call failed"
                        processed_df.at[index, '_temp_conversion_status'] = error_message
                    
                except Exception as e:
                    failed_calls += 2
                    failed_properties.append(index)
                    logger.error(f"   ❌ Failed to analyze property {index + 1}: {str(e)} - INCLUDING in output with error message")
                    
                    # Store zeros for failed properties (keep original input data intact)
                    processed_df.at[index, 'Vegetation (NDVI)-Before Value'] = 0.0
                    processed_df.at[index, 'Vegetation (NDVI)-After Value'] = 0.0
                    processed_df.at[index, 'Built-up Area (NDBI)-Before Value'] = 0.0
                    processed_df.at[index, 'Built-up Area (NDBI)-After Value'] = 0.0
                    processed_df.at[index, 'Water/Moisture (NDWI)-Before Value'] = 0.0
                    processed_df.at[index, 'Water/Moisture (NDWI)-After Value'] = 0.0
                    
                    # Store exception message for later (will be moved to end)
                    processed_df.at[index, '_temp_conversion_status'] = str(e)
            
            # Keep all properties in output (no longer removing failed ones)
            logger.info(f"📊 Keeping all {len(processed_df)} properties in output (including failed ones with error messages)")
            if failed_properties:
                logger.info(f"   Failed property indices: {failed_properties}")
            
            # Log batch processing summary
            self._log_batch_summary(len(df), successful_calls, failed_calls)
            
            # Additional summary for successful vs failed properties
            logger.info(f"📋 PROPERTY PROCESSING SUMMARY")
            logger.info(f"   Total Properties Attempted: {len(df)}")
            logger.info(f"   Successfully Processed: {len(successful_properties)}")
            logger.info(f"   Failed (with error messages): {len(failed_properties)}")
            logger.info(f"   Properties in Output: {len(processed_df)}")
            
            return processed_df, successful_calls
            
        except Exception as e:
            raise FileProcessingException(f"Failed to get real satellite indices: {str(e)}")
    
    async def _analyze_single_property(self, lat: float, lon: float, extent_ac: float, 
                                     time_period: Tuple[str, str]) -> Tuple[Dict[str, float], str]:
        """Analyze a single property using Sentinel Hub API with comprehensive logging.
        
        Returns:
            Tuple[Dict[str, float], str]: (indices_dict, error_message)
            - indices_dict: Dictionary with ndvi, ndbi, ndwi values
            - error_message: Empty string if successful, error message if failed
        """
        
        # Generate unique request ID for tracking
        request_id = f"SH_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{self.api_call_count + 1:03d}"
        start_time = time.time()
        
        try:
            # Create bounding box considering land area
            bbox = self._create_property_bbox(lat, lon, extent_ac)
            size = bbox_to_dimensions(bbox, resolution=self.resolution)
            
            # Log request details
            self._log_api_request(request_id, lat, lon, time_period, bbox, size)
            
            # Evalscript for calculating indices
            evalscript = self.config_data['evalscripts']['change_detection']
            
            # Log evalscript being used
            logger.debug(f"   Using evalscript: change_detection")
            logger.debug(f"   Evalscript length: {len(evalscript)} characters")
            
            # Create Sentinel Hub request
            request = SentinelHubRequest(
                evalscript=evalscript,
                input_data=[
                    SentinelHubRequest.input_data(
                        data_collection=DataCollection.SENTINEL2_L2A,
                        time_interval=time_period,
                        maxcc=self.max_cloud_coverage/100.0
                    )
                ],
                responses=[
                    SentinelHubRequest.output_response('default', MimeType.TIFF)
                ],
                bbox=bbox,
                size=size,
                config=self.config
            )
            
            # Log request configuration
            logger.debug(f"   Request URL: {request.base_url if hasattr(request, 'base_url') else 'N/A'}")
            logger.debug(f"   OAuth Token: {'Present' if self.config.sh_client_id else 'Missing'}")
            
            # Make API call and measure time
            api_start_time = time.time()
            logger.info(f"   🔄 Making API call to Sentinel Hub...")
            
            response = request.get_data()
            
            api_end_time = time.time()
            response_time = api_end_time - start_time
            
            if response and len(response) > 0:
                data = response[0]
                data_shape = data.shape
                
                # Calculate mean values for each index
                indices = {
                    'ndvi': float(np.mean(data[:, :, 0])),
                    'ndbi': float(np.mean(data[:, :, 1])),
                    'ndwi': float(np.mean(data[:, :, 2]))
                }
                
                # Log successful response
                self._log_api_response(request_id, True, response_time, data_shape, None, indices)
                
                # Additional data quality logging
                logger.debug(f"   Data Quality Check:")
                logger.debug(f"     NDVI range: {np.min(data[:, :, 0]):.4f} to {np.max(data[:, :, 0]):.4f}")
                logger.debug(f"     NDBI range: {np.min(data[:, :, 1]):.4f} to {np.max(data[:, :, 1]):.4f}")
                logger.debug(f"     NDWI range: {np.min(data[:, :, 2]):.4f} to {np.max(data[:, :, 2]):.4f}")
                
                return indices, ""  # Success - no error message
            else:
                # Log no data response
                error_msg = f"No satellite data available for period {time_period}"
                self._log_api_response(request_id, False, response_time, None, error_msg)
                return {'ndvi': 0.0, 'ndbi': 0.0, 'ndwi': 0.0}, error_msg
                
        except Exception as e:
            response_time = time.time() - start_time
            error_msg = str(e)
            
            # Extract more specific error message from Sentinel Hub response
            if "Server response:" in error_msg:
                try:
                    # Extract the JSON error message from Sentinel Hub
                    import re
                    json_match = re.search(r'Server response: "({.*?})"', error_msg)
                    if json_match:
                        import json
                        error_json = json.loads(json_match.group(1))
                        specific_error = error_json.get('message', error_msg)
                        error_msg = specific_error
                except:
                    pass  # Use original error message if parsing fails
            
            # Log failed response
            self._log_api_response(request_id, False, response_time, None, error_msg)
            
            # Log additional error details
            logger.error(f"   Exception Type: {type(e).__name__}")
            logger.error(f"   Exception Details: {error_msg}")
            
            # Return default values and error message on API failure
            return {'ndvi': 0.0, 'ndbi': 0.0, 'ndwi': 0.0}, error_msg
    
    def _log_api_request(self, request_id: str, lat: float, lon: float, 
                        time_period: Tuple[str, str], bbox: BBox, size: Tuple[int, int]):
        """Log detailed information about Sentinel Hub API request."""
        logger.info(f"🛰️ SENTINEL HUB API REQUEST #{self.api_call_count + 1}")
        logger.info(f"   Request ID: {request_id}")
        logger.info(f"   Location: {lat:.6f}, {lon:.6f}")
        logger.info(f"   Time Period: {time_period[0]} to {time_period[1]}")
        logger.info(f"   Bounding Box: [{bbox.lower_left[0]:.6f}, {bbox.lower_left[1]:.6f}, {bbox.upper_right[0]:.6f}, {bbox.upper_right[1]:.6f}]")
        logger.info(f"   Image Size: {size[0]}x{size[1]} pixels")
        logger.info(f"   Resolution: {self.resolution}m per pixel")
        logger.info(f"   Max Cloud Coverage: {self.max_cloud_coverage}%")
        logger.info(f"   Data Collection: SENTINEL2_L2A")
    
    def _log_api_response(self, request_id: str, success: bool, response_time: float, 
                         data_shape: Tuple = None, error: str = None, 
                         indices: Dict[str, float] = None):
        """Log detailed information about Sentinel Hub API response."""
        status = "SUCCESS" if success else "FAILED"
        logger.info(f"📡 SENTINEL HUB API RESPONSE #{self.api_call_count}")
        logger.info(f"   Request ID: {request_id}")
        logger.info(f"   Status: {status}")
        logger.info(f"   Response Time: {response_time:.3f} seconds")
        
        if success and data_shape:
            logger.info(f"   Data Shape: {data_shape}")
            if indices:
                logger.info(f"   NDVI: {indices['ndvi']:.6f}")
                logger.info(f"   NDBI: {indices['ndbi']:.6f}")
                logger.info(f"   NDWI: {indices['ndwi']:.6f}")
        elif error:
            logger.error(f"   Error: {error}")
        
        # Update statistics
        self.api_call_count += 1
        self.total_api_time += response_time
        avg_time = self.total_api_time / self.api_call_count
        logger.info(f"   API Stats: {self.api_call_count} calls, {self.total_api_time:.1f}s total, {avg_time:.2f}s avg")
    
    def _log_batch_summary(self, total_properties: int, successful_calls: int, failed_calls: int):
        """Log summary of batch processing."""
        logger.info(f"📊 SENTINEL HUB BATCH PROCESSING SUMMARY")
        logger.info(f"   Total Properties: {total_properties}")
        logger.info(f"   Successful API Calls: {successful_calls}")
        logger.info(f"   Failed API Calls: {failed_calls}")
        logger.info(f"   Success Rate: {(successful_calls/(successful_calls+failed_calls)*100):.1f}%")
        logger.info(f"   Total API Time: {self.total_api_time:.1f} seconds")
        logger.info(f"   Average Time per Call: {(self.total_api_time/self.api_call_count):.2f} seconds")
    
    def _create_property_bbox(self, lat: float, lon: float, extent_ac: float = 0, 
                            buffer_meters: float = 50) -> BBox:
        """Create bounding box for a property with buffer, considering land area."""
        
        # Convert acres to square meters (1 acre = 4046.86 sq meters)
        extent_sq_meters = extent_ac * 4046.86 if extent_ac > 0 else 0
        
        # Calculate property radius based on area (assuming circular property)
        if extent_sq_meters > 0:
            property_radius_meters = (extent_sq_meters / 3.14159) ** 0.5
        else:
            property_radius_meters = 50  # Default 50m radius for point locations
        
        # Use the larger of property radius or buffer
        effective_radius_meters = max(property_radius_meters, buffer_meters)
        
        # Convert radius from meters to degrees (approximate)
        radius_degrees = effective_radius_meters / 111000  # 1 degree ≈ 111km
        
        bbox = BBox(
            bbox=[
                lon - radius_degrees,
                lat - radius_degrees,
                lon + radius_degrees,
                lat + radius_degrees
            ],
            crs=CRS.WGS84
        )
        
        return bbox
    
    async def _calculate_differences_and_interpretations(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calculate differences and provide interpretations for each index, but only for successful properties."""
        
        try:
            processed_df = df.copy()
            
            # Initialize all difference and interpretation columns with empty/default values
            processed_df['Vegetation (NDVI)-Difference'] = ''
            processed_df['Vegetation (NDVI)-Interpretation'] = ''
            processed_df['Vegetation (NDVI)-Significance'] = ''
            
            processed_df['Built-up Area (NDBI)-Difference'] = ''
            processed_df['Built-up Area (NDBI)-Interpretation'] = ''
            processed_df['Built-up Area (NDBI)-Significance'] = ''
            
            processed_df['Water/Moisture (NDWI)-Difference'] = ''
            processed_df['Water/Moisture (NDWI)-Interpretation'] = ''
            processed_df['Water/Moisture (NDWI)-Significance'] = ''
            
            # Only calculate differences and interpretations for successful properties
            for index, row in processed_df.iterrows():
                if '_temp_conversion_status' in row and row['_temp_conversion_status'] == 'Successful':
                    # Calculate NDVI differences and interpretations
                    ndvi_diff = round(row['Vegetation (NDVI)-After Value'] - row['Vegetation (NDVI)-Before Value'], 4)
                    processed_df.at[index, 'Vegetation (NDVI)-Difference'] = ndvi_diff
                    processed_df.at[index, 'Vegetation (NDVI)-Interpretation'] = self._interpret_ndvi_change(ndvi_diff)
                    processed_df.at[index, 'Vegetation (NDVI)-Significance'] = 'Yes' if abs(ndvi_diff) >= self.thresholds['default_threshold'] else 'No'
                    
                    # Calculate NDBI differences and interpretations
                    ndbi_diff = round(row['Built-up Area (NDBI)-After Value'] - row['Built-up Area (NDBI)-Before Value'], 4)
                    processed_df.at[index, 'Built-up Area (NDBI)-Difference'] = ndbi_diff
                    processed_df.at[index, 'Built-up Area (NDBI)-Interpretation'] = self._interpret_ndbi_change(ndbi_diff)
                    processed_df.at[index, 'Built-up Area (NDBI)-Significance'] = 'Yes' if abs(ndbi_diff) >= self.thresholds['default_threshold'] else 'No'
                    
                    # Calculate NDWI differences and interpretations
                    ndwi_diff = round(row['Water/Moisture (NDWI)-After Value'] - row['Water/Moisture (NDWI)-Before Value'], 4)
                    processed_df.at[index, 'Water/Moisture (NDWI)-Difference'] = ndwi_diff
                    processed_df.at[index, 'Water/Moisture (NDWI)-Interpretation'] = self._interpret_ndwi_change(ndwi_diff)
                    processed_df.at[index, 'Water/Moisture (NDWI)-Significance'] = 'Yes' if abs(ndwi_diff) >= self.thresholds['ndwi_thresholds']['water_appearance'] else 'No'
                else:
                    # For failed properties, leave ALL analysis fields as empty spaces
                    processed_df.at[index, 'Vegetation (NDVI)-Difference'] = ''
                    processed_df.at[index, 'Vegetation (NDVI)-Interpretation'] = ''
                    processed_df.at[index, 'Vegetation (NDVI)-Significance'] = ''
                    
                    processed_df.at[index, 'Built-up Area (NDBI)-Difference'] = ''
                    processed_df.at[index, 'Built-up Area (NDBI)-Interpretation'] = ''
                    processed_df.at[index, 'Built-up Area (NDBI)-Significance'] = ''
                    
                    processed_df.at[index, 'Water/Moisture (NDWI)-Difference'] = ''
                    processed_df.at[index, 'Water/Moisture (NDWI)-Interpretation'] = ''
                    processed_df.at[index, 'Water/Moisture (NDWI)-Significance'] = ''
            
            # Move Conversion_status to the end (last column)
            if '_temp_conversion_status' in processed_df.columns:
                processed_df['Conversion_status'] = processed_df['_temp_conversion_status']
                processed_df = processed_df.drop('_temp_conversion_status', axis=1)
            
            return processed_df
            
        except Exception as e:
            raise FileProcessingException(f"Failed to calculate differences and interpretations: {str(e)}")
    
    def _interpret_ndvi_change(self, difference: float) -> str:
        """Interpret NDVI change based on difference value."""
        if difference >= self.thresholds['ndvi_thresholds']['moderate_increase']:
            return "Vegetation growth or improvement"
        elif difference <= self.thresholds['ndvi_thresholds']['moderate_decrease']:
            return "Vegetation loss or degradation"
        else:
            return "No significant vegetation change"
    
    def _interpret_ndbi_change(self, difference: float) -> str:
        """Interpret NDBI change based on difference value."""
        if difference >= self.thresholds['ndbi_thresholds']['minor_increase']:
            return "Construction or development increase"
        elif difference <= self.thresholds['ndbi_thresholds']['demolition']:
            return "Construction or development decrease"
        else:
            return "No significant built-up area change"
    
    def _interpret_ndwi_change(self, difference: float) -> str:
        """Interpret NDWI change based on difference value."""
        if difference >= self.thresholds['ndwi_thresholds']['water_appearance']:
            return "Water increase or flooding"
        elif difference <= self.thresholds['ndwi_thresholds']['water_reduction']:
            return "Water decrease or drought"
        else:
            return "No significant water change"
    
    async def _generate_output_file(
        self, 
        df: pd.DataFrame, 
        output_dir: Path, 
        original_filename: str,
        engagement_name: str
    ) -> Path:
        """Generate the processed output file with proper status formatting."""
        
        try:
            # Create output filename with timestamp and engagement info
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            before_period = df['Before Period Start'].iloc[0].replace('-', '').replace(' ', '') if len(df) > 0 else 'before'
            after_period = df['After Period Start'].iloc[0].replace('-', '').replace(' ', '') if len(df) > 0 else 'after'
            
            # Generate CSV output (primary format)
            csv_filename = f"{timestamp}_batch_analysis_before{before_period}_{after_period}.csv"
            csv_path = output_dir / csv_filename
            df.to_csv(csv_path, index=False)
            
            # Generate Excel output with color formatting
            excel_filename = f"{timestamp}_batch_analysis_before{before_period}_{after_period}.xlsx"
            excel_path = output_dir / excel_filename
            self._generate_excel_output(df, excel_path, engagement_name)
            
            # Generate HTML output with colored status column
            html_filename = f"{timestamp}_batch_analysis_before{before_period}_{after_period}.html"
            html_path = output_dir / html_filename
            self._generate_html_output(df, html_path, engagement_name)
            
            logger.info(f"Real Sentinel Hub output files generated:")
            logger.info(f"  CSV: {csv_path}")
            logger.info(f"  Excel: {excel_path}")
            logger.info(f"  HTML: {html_path}")
            
            return csv_path  # Return CSV path as primary output
            
        except Exception as e:
            raise FileProcessingException(f"Failed to generate output file: {str(e)}")
    
    def _generate_excel_output(self, df: pd.DataFrame, excel_path: Path, engagement_name: str):
        """Generate Excel output with color formatting for successful/failed rows."""
        
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            
            # Write DataFrame to Excel
            df.to_excel(excel_path, index=False, engine='openpyxl')
            
            # Load the workbook and worksheet for formatting
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active
            
            # Define colors
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            green_font = Font(color='008000', bold=True)  # Green font
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            red_font = Font(color='FF0000')  # Red font
            
            # Find the Conversion_status column index
            conversion_status_col = None
            for col_idx, cell in enumerate(worksheet[1], 1):  # Header row
                if cell.value == 'Conversion_status':
                    conversion_status_col = col_idx
                    break
            
            if conversion_status_col:
                # Apply formatting to each data row
                for row_idx in range(2, worksheet.max_row + 1):  # Skip header row
                    status_cell = worksheet.cell(row=row_idx, column=conversion_status_col)
                    status_value = status_cell.value
                    
                    if status_value == 'Successful':
                        # Green background and font for successful status cell only
                        status_cell.fill = green_fill
                        status_cell.font = green_font
                    else:
                        # Red background and font for entire failed row
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = red_fill
                            cell.font = red_font
            
            # Save the formatted workbook
            workbook.save(excel_path)
            logger.info(f"Excel file with color formatting saved: {excel_path}")
            
        except ImportError:
            logger.warning("openpyxl not available - generating Excel without color formatting")
            df.to_excel(excel_path, index=False)
        except Exception as e:
            logger.warning(f"Failed to generate colored Excel output: {str(e)} - generating basic Excel")
            df.to_excel(excel_path, index=False)
    
    def _generate_html_output(self, df: pd.DataFrame, html_path: Path, engagement_name: str):
        """Generate HTML output with colored rows and status cells."""
        
        try:
            # Create a copy of the dataframe for HTML formatting
            html_df = df.copy()
            
            # Apply color formatting to entire rows based on Conversion_status
            def color_row(row):
                if 'Conversion_status' in row:
                    status = row['Conversion_status']
                    if status == 'Successful':
                        # Green background for successful status cell only
                        styles = [''] * len(row)
                        status_col_idx = row.index.get_loc('Conversion_status')
                        styles[status_col_idx] = 'background-color: #ccffcc; color: green; font-weight: bold;'
                        return styles
                    else:
                        # Red background for entire failed row
                        return ['background-color: #ffcccc; color: red;'] * len(row)
                return [''] * len(row)
            
            # Create styled HTML
            styled_df = html_df.style.apply(color_row, axis=1)
            
            # Generate HTML with custom styling
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Sentinel Hub Analysis Results - {engagement_name}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #333; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f2f2f2; font-weight: bold; }}
                    .failed {{ background-color: #ffcccc !important; color: red !important; font-weight: bold; }}
                    .success {{ background-color: #ccffcc !important; color: green !important; font-weight: bold; }}
                </style>
            </head>
            <body>
                <h1>Sentinel Hub Satellite Analysis Results</h1>
                <p><strong>Engagement:</strong> {engagement_name}</p>
                <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p><strong>Total Properties:</strong> {len(html_df)}</p>
                
                {styled_df.to_html(escape=False, table_id='results_table')}
                
                <script>
                    // Additional JavaScript to ensure proper row coloring
                    document.addEventListener('DOMContentLoaded', function() {{
                        const rows = document.querySelectorAll('tr');
                        rows.forEach(row => {{
                            const statusCell = row.querySelector('td:last-child'); // Assuming Conversion_status is last column
                            if (statusCell) {{
                                const status = statusCell.textContent.trim();
                                if (status === 'Successful') {{
                                    statusCell.style.backgroundColor = '#ccffcc';
                                    statusCell.style.color = 'green';
                                    statusCell.style.fontWeight = 'bold';
                                }} else if (status !== 'Successful' && status !== '') {{
                                    // Color entire row red for failed cases
                                    const cells = row.querySelectorAll('td');
                                    cells.forEach(cell => {{
                                        cell.style.backgroundColor = '#ffcccc';
                                        cell.style.color = 'red';
                                    }});
                                }}
                            }}
                        }});
                    }});
                </script>
            </body>
            </html>
            """
            
            # Write HTML file
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
        except Exception as e:
            logger.warning(f"Failed to generate HTML output: {str(e)}")
            # Don't raise exception as CSV is the primary output