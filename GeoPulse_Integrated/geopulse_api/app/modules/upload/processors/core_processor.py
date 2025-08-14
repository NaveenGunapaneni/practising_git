"""Core file processor for generic business data analysis.

This processor handles non-geospatial data and provides:
- Statistical analysis (sum, average, variance, etc.)
- Quality scoring and categorization
- Excel output with conditional formatting
- Summary sheets with key metrics

For geospatial data (with coordinates), use GeospatialProcessor instead.
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

from app.core.exceptions import FileProcessingException
from app.core.logger import get_logger

logger = get_logger(__name__)


class CoreFileProcessor:
    """Service for processing uploaded files with core business logic."""
    
    def __init__(self):
        self.supported_formats = {'.csv', '.xlsx', '.xls'}
    
    async def process_file(
        self,
        input_path: Path,
        output_dir: Path,
        dates: List[str],
        engagement_name: str
    ) -> Path:
        """Process the uploaded file with core business logic."""
        
        logger.info(f"Starting file processing: {input_path}")
        
        try:
            # Step 1: Load the file
            df = await self._load_file(input_path)
            
            # Step 2: Apply core business logic
            processed_df = await self._apply_business_logic(df, dates, engagement_name)
            
            # Step 3: Apply conditional formatting and enhancements
            formatted_df = await self._apply_conditional_formatting(processed_df)
            
            # Step 4: Generate output file
            output_path = await self._generate_output_file(
                formatted_df, 
                output_dir, 
                input_path.stem,
                engagement_name
            )
            
            logger.info(f"File processing completed: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"File processing failed for {input_path}: {str(e)}")
            raise FileProcessingException(f"Processing failed: {str(e)}")
    
    async def _load_file(self, file_path: Path) -> pd.DataFrame:
        """Load file into pandas DataFrame."""
        
        try:
            file_extension = file_path.suffix.lower()
            
            if file_extension == '.csv':
                # Try different encodings and separators for CSV
                for encoding in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        for sep in [',', ';', '\t']:
                            try:
                                df = pd.read_csv(file_path, encoding=encoding, sep=sep)
                                if len(df.columns) > 1:  # Valid CSV should have multiple columns
                                    logger.info(f"CSV loaded with encoding={encoding}, sep='{sep}'")
                                    return df
                            except:
                                continue
                    except:
                        continue
                
                # Fallback to default CSV reading
                df = pd.read_csv(file_path)
                
            elif file_extension in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path)
                
            else:
                raise FileProcessingException(f"Unsupported file format: {file_extension}")
            
            if df.empty:
                raise FileProcessingException("File is empty or contains no data")
            
            logger.info(f"File loaded successfully: {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except Exception as e:
            raise FileProcessingException(f"Failed to load file: {str(e)}")
    
    async def _apply_business_logic(
        self, 
        df: pd.DataFrame, 
        dates: List[str], 
        engagement_name: str
    ) -> pd.DataFrame:
        """Apply core business logic transformations."""
        
        logger.info("Applying core business logic transformations")
        
        try:
            # Create a copy to avoid modifying original data
            processed_df = df.copy()
            
            # Add metadata columns
            processed_df['engagement_name'] = engagement_name
            processed_df['processing_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Add the four dates as separate columns
            for i, date in enumerate(dates, 1):
                processed_df[f'date_{i}'] = date
            
            # Core business logic transformations
            processed_df = await self._apply_data_transformations(processed_df)
            processed_df = await self._apply_calculations(processed_df)
            processed_df = await self._apply_aggregations(processed_df)
            processed_df = await self._apply_inference_logic(processed_df)
            
            logger.info(f"Business logic applied: {len(processed_df)} rows processed")
            return processed_df
            
        except Exception as e:
            raise FileProcessingException(f"Business logic application failed: {str(e)}")
    
    async def _apply_data_transformations(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply data transformation rules."""
        
        try:
            # Clean and standardize data
            for col in df.select_dtypes(include=['object']).columns:
                if col not in ['engagement_name', 'processing_date']:
                    # Strip whitespace and standardize text
                    df[col] = df[col].astype(str).str.strip()
                    
                    # Convert empty strings to NaN
                    df[col] = df[col].replace('', pd.NA)
            
            # Handle numeric columns
            for col in df.select_dtypes(include=['number']).columns:
                # Fill NaN values with 0 for numeric columns (business rule)
                df[col] = df[col].fillna(0)
            
            # Add row index for tracking
            df['row_index'] = range(1, len(df) + 1)
            
            return df
            
        except Exception as e:
            logger.warning(f"Data transformation warning: {str(e)}")
            return df
    
    async def _apply_calculations(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply mathematical calculations and business rules."""
        
        try:
            # Find numeric columns for calculations
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
            
            # Remove metadata columns from calculations
            exclude_cols = ['row_index']
            numeric_cols = [col for col in numeric_cols if col not in exclude_cols]
            
            if len(numeric_cols) >= 2:
                # Example business calculations
                df['total_sum'] = df[numeric_cols].sum(axis=1)
                df['average'] = df[numeric_cols].mean(axis=1)
                df['max_value'] = df[numeric_cols].max(axis=1)
                df['min_value'] = df[numeric_cols].min(axis=1)
                
                # Calculate variance and standard deviation
                df['variance'] = df[numeric_cols].var(axis=1)
                df['std_deviation'] = df[numeric_cols].std(axis=1)
                
                # Business rule: Flag high variance rows
                if 'variance' in df.columns:
                    variance_threshold = df['variance'].quantile(0.75)  # Top 25%
                    df['high_variance_flag'] = df['variance'] > variance_threshold
            
            # Add calculated fields based on business requirements
            df['processing_score'] = self._calculate_processing_score(df)
            
            return df
            
        except Exception as e:
            logger.warning(f"Calculation warning: {str(e)}")
            return df
    
    def _calculate_processing_score(self, df: pd.DataFrame) -> pd.Series:
        """Calculate a processing score based on business logic."""
        
        try:
            # Initialize score
            score = pd.Series(50, index=df.index)  # Base score of 50
            
            # Adjust score based on data completeness
            for col in df.select_dtypes(include=['object']).columns:
                if col not in ['engagement_name', 'processing_date']:
                    # Add points for non-null values
                    score += (~df[col].isna()).astype(int) * 5
            
            # Adjust score based on numeric values
            numeric_cols = df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                if col not in ['row_index', 'total_sum', 'average', 'max_value', 'min_value', 'variance', 'std_deviation']:
                    # Add points for positive values
                    score += (df[col] > 0).astype(int) * 3
            
            # Cap score at 100
            score = score.clip(upper=100)
            
            return score
            
        except Exception as e:
            logger.warning(f"Score calculation warning: {str(e)}")
            return pd.Series(50, index=df.index)
    
    async def _apply_aggregations(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply data aggregations and grouping logic."""
        
        try:
            # Add summary statistics as new columns
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
            
            # Remove calculated columns from aggregation
            exclude_cols = ['row_index', 'total_sum', 'average', 'max_value', 'min_value', 'variance', 'std_deviation', 'processing_score']
            base_numeric_cols = [col for col in numeric_cols if col not in exclude_cols]
            
            if base_numeric_cols:
                # Add percentile rankings
                for col in base_numeric_cols:
                    df[f'{col}_percentile'] = df[col].rank(pct=True) * 100
                
                # Add quartile classifications
                for col in base_numeric_cols:
                    df[f'{col}_quartile'] = pd.qcut(df[col], q=4, labels=['Q1', 'Q2', 'Q3', 'Q4'], duplicates='drop')
            
            return df
            
        except Exception as e:
            logger.warning(f"Aggregation warning: {str(e)}")
            return df
    
    async def _apply_inference_logic(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply inference and machine learning logic."""
        
        try:
            # Business rule-based classifications
            if 'processing_score' in df.columns:
                # Classify records based on processing score
                df['quality_category'] = pd.cut(
                    df['processing_score'],
                    bins=[0, 30, 60, 80, 100],
                    labels=['Poor', 'Fair', 'Good', 'Excellent'],
                    include_lowest=True
                )
            
            # Add recommendation flags
            df['requires_review'] = False
            
            # Flag records that require review based on business rules
            if 'high_variance_flag' in df.columns:
                df.loc[df['high_variance_flag'] == True, 'requires_review'] = True
            
            if 'processing_score' in df.columns:
                df.loc[df['processing_score'] < 40, 'requires_review'] = True
            
            # Add priority scoring
            df['priority_score'] = 1  # Default priority
            df.loc[df['requires_review'] == True, 'priority_score'] = 3  # High priority
            
            return df
            
        except Exception as e:
            logger.warning(f"Inference logic warning: {str(e)}")
            return df
    
    async def _apply_conditional_formatting(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply conditional formatting and styling rules."""
        
        try:
            # Add formatting indicators for Excel output
            df['format_style'] = 'normal'
            
            # Apply conditional formatting rules
            if 'requires_review' in df.columns:
                df.loc[df['requires_review'] == True, 'format_style'] = 'highlight_red'
            
            if 'quality_category' in df.columns:
                df.loc[df['quality_category'] == 'Excellent', 'format_style'] = 'highlight_green'
                df.loc[df['quality_category'] == 'Poor', 'format_style'] = 'highlight_red'
            
            # Sort by priority and processing score
            if 'priority_score' in df.columns and 'processing_score' in df.columns:
                df = df.sort_values(['priority_score', 'processing_score'], ascending=[False, False])
            
            return df
            
        except Exception as e:
            logger.warning(f"Conditional formatting warning: {str(e)}")
            return df
    
    async def _generate_output_file(
        self, 
        df: pd.DataFrame, 
        output_dir: Path, 
        original_filename: str,
        engagement_name: str
    ) -> Path:
        """Generate the processed output file."""
        
        try:
            # Create output filename
            timestamp = datetime.now().strftime("%H%M%S")
            output_filename = f"processed_{timestamp}_{original_filename}.xlsx"
            output_path = output_dir / output_filename
            
            # Write to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Processed_Data', index=False)
                
                # Write summary sheet
                summary_df = self._create_summary_sheet(df, engagement_name)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Apply Excel formatting
                self._apply_excel_formatting(writer, df)
            
            logger.info(f"Output file generated: {output_path}")
            return output_path
            
        except Exception as e:
            raise FileProcessingException(f"Failed to generate output file: {str(e)}")
    
    def _create_summary_sheet(self, df: pd.DataFrame, engagement_name: str) -> pd.DataFrame:
        """Create a summary sheet with key metrics."""
        
        try:
            summary_data = {
                'Metric': [],
                'Value': []
            }
            
            # Basic metrics
            summary_data['Metric'].extend([
                'Engagement Name',
                'Processing Date',
                'Total Records',
                'Records Requiring Review',
                'Average Processing Score'
            ])
            
            summary_data['Value'].extend([
                engagement_name,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                len(df),
                len(df[df.get('requires_review', False) == True]) if 'requires_review' in df.columns else 0,
                f"{df['processing_score'].mean():.2f}" if 'processing_score' in df.columns else 'N/A'
            ])
            
            # Quality distribution
            if 'quality_category' in df.columns:
                quality_counts = df['quality_category'].value_counts()
                for category, count in quality_counts.items():
                    summary_data['Metric'].append(f'Records - {category}')
                    summary_data['Value'].append(count)
            
            return pd.DataFrame(summary_data)
            
        except Exception as e:
            logger.warning(f"Summary sheet creation warning: {str(e)}")
            return pd.DataFrame({'Metric': ['Error'], 'Value': ['Failed to create summary']})
    
    def _apply_excel_formatting(self, writer, df: pd.DataFrame):
        """Apply Excel formatting to the output file."""
        
        try:
            from openpyxl.styles import PatternFill, Font
            
            workbook = writer.book
            worksheet = writer.sheets['Processed_Data']
            
            # Define colors
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            header_fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
            bold_font = Font(bold=True)
            
            # Format headers
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = bold_font
            
            # Apply conditional formatting based on format_style column
            if 'format_style' in df.columns:
                style_col_idx = df.columns.get_loc('format_style') + 1
                
                for row in range(2, len(df) + 2):  # Skip header row
                    style_value = worksheet.cell(row=row, column=style_col_idx).value
                    
                    if style_value == 'highlight_red':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = red_fill
                    elif style_value == 'highlight_green':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = green_fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.warning(f"Excel formatting warning: {str(e)}")
            # Continue without formatting if there's an error